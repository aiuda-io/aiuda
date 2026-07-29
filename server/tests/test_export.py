"""Exportar a Excel (GET /v1/export/{entidad}.xlsx): el archivo es xlsx válido con
encabezados reimportables, respeta los filtros que el usuario ve en la página,
aísla por tenant y deja bitácora (data.export).

El test redondo: un Excel exportado por aiuda pasa por el importador
(smart_import.commit con mapeo identidad — los encabezados SON los campos que el
importador entiende) y las mismas facturas se reconocen: se crean idénticas en un
negocio limpio y se saltan como duplicadas en el negocio de origen."""

import io
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from aiuda_server.api.main import app, get_db
from aiuda_core.config import settings
from aiuda_core.connectors.smart_import import ENTITY_FIELDS, commit, read_table
from aiuda_core.models import (
    Appointment,
    AuditLog,
    Base,
    Customer,
    Invoice,
    Payment,
    PaymentPromise,
    Product,
    Tenant,
)

HEADERS = {"X-API-Key": "k-demo"}
HOY = datetime.now(ZoneInfo("America/Mexico_City")).date()

FACTURAS_HDR = [
    "folio", "cliente", "telefono", "monto", "fecha_emision", "fecha_vencimiento",
    "estado", "tramo", "procedencia", "pagada_el",
]


@pytest.fixture()
def db_session():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
    session = SessionLocal()
    yield session
    session.close()


@pytest.fixture()
def client(db_session):
    app.dependency_overrides[get_db] = lambda: db_session
    app.state.queue = None
    yield TestClient(app)
    app.dependency_overrides.clear()


@pytest.fixture()
def tenant(db_session):
    t = Tenant(
        name="Demo SA",
        owner_phone="5215512345678",
        evolution_instance="demo",
        config={
            "api_key": "k-demo",
            "tags": [{"id": "tg1", "name": "Mayoreo", "color": "#0a0"}],
        },
    )
    db_session.add(t)
    db_session.flush()
    return t


@pytest.fixture()
def otro_tenant(db_session):
    t = Tenant(
        name="Otro SA",
        owner_phone="5215500000000",
        evolution_instance="otro",
        config={"api_key": "k-otro"},
    )
    db_session.add(t)
    db_session.flush()
    return t


def _cust(db, tenant, name, phone, **kw):
    c = Customer(tenant_id=tenant.id, name=name, phone=phone, **kw)
    db.add(c)
    db.flush()
    return c


def _inv(db, tenant, customer, folio, amount, due, **kw):
    inv = Invoice(
        tenant_id=tenant.id, customer_id=customer.id, folio=folio, amount=amount,
        issued_date=due - timedelta(days=30), due_date=due, **kw,
    )
    db.add(inv)
    db.flush()
    return inv


def _sheet(res):
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return ws, headers, rows


def _seed_facturas(db, tenant):
    ana = _cust(db, tenant, "Papelería Ana", "5215511110001")
    beto = _cust(db, tenant, "Ferretería Beto", "5215511110002")
    # Tramos distintos: crítica (vencida hace 60 días) y por vencer (en 10 días).
    critica = _inv(db, tenant, ana, "F-001", 1500.50, HOY - timedelta(days=60))
    porvencer = _inv(db, tenant, beto, "F-002", 800.00, HOY + timedelta(days=10))
    pagada = _inv(
        db, tenant, ana, "F-003", 999.99, HOY - timedelta(days=5),
        status="paid", paid_at=datetime(2026, 6, 1, 12, 0),
    )
    return critica, porvencer, pagada


# --- Forma del archivo --------------------------------------------------------


def test_facturas_xlsx_valido(client, db_session, tenant):
    _seed_facturas(db_session, tenant)
    res = client.get("/v1/export/facturas.xlsx", headers=HEADERS)
    assert res.status_code == 200
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    dispo = res.headers["content-disposition"]
    assert "attachment" in dispo and f'facturas-{HOY.isoformat()}.xlsx' in dispo

    ws, headers, rows = _sheet(res)
    assert ws.title == "facturas"
    assert headers == FACTURAS_HDR
    # Encabezado en negritas y congelado en A2 (nada más de estilo).
    assert ws["A1"].font.bold is True
    assert ws.freeze_panes == "A2"
    # Default = abiertas (como la pestaña): la pagada no aparece.
    assert len(rows) == 2
    fila = {r[0]: r for r in rows}["F-001"]
    assert fila[1] == "Papelería Ana"
    assert fila[2] == "5215511110001"
    assert isinstance(fila[3], float) and fila[3] == 1500.50  # monto como número
    assert isinstance(fila[5], datetime)  # fecha como fecha, no string
    assert fila[5].date() == HOY - timedelta(days=60)
    assert fila[6] == "abierta" and fila[7] == "Crítica +45 días" and fila[8] == "excel"


def test_facturas_respeta_filtros(client, db_session, tenant):
    _seed_facturas(db_session, tenant)
    # status=paid: solo la pagada, con su fecha de pago.
    _, _, rows = _sheet(client.get("/v1/export/facturas.xlsx?status=paid", headers=HEADERS))
    assert [r[0] for r in rows] == ["F-003"]
    assert rows[0][6] == "pagada" and rows[0][9] is not None
    # q filtra por cliente o folio (igual que la búsqueda de la página).
    _, _, rows = _sheet(client.get("/v1/export/facturas.xlsx?q=beto", headers=HEADERS))
    assert [r[0] for r in rows] == ["F-002"]
    _, _, rows = _sheet(client.get("/v1/export/facturas.xlsx?q=F-001", headers=HEADERS))
    assert [r[0] for r in rows] == ["F-001"]
    # bucket filtra por tramo de cartera.
    _, _, rows = _sheet(client.get("/v1/export/facturas.xlsx?bucket=critica", headers=HEADERS))
    assert [r[0] for r in rows] == ["F-001"]


def test_facturas_aisla_tenant(client, db_session, tenant, otro_tenant):
    _seed_facturas(db_session, tenant)
    ajeno = _cust(db_session, otro_tenant, "Cliente Ajeno", "5215599999999")
    _inv(db_session, otro_tenant, ajeno, "X-666", 5000, HOY)

    _, _, rows = _sheet(client.get("/v1/export/facturas.xlsx", headers=HEADERS))
    assert {r[0] for r in rows} == {"F-001", "F-002"}
    # Cambiar el workspace activo cambia QUÉ datos salen: el filtro por tenant
    # sigue vivo aunque ya no haya sesiones.
    settings.workspace_id = otro_tenant.id
    _, _, rows = _sheet(client.get("/v1/export/facturas.xlsx", headers=HEADERS))
    assert {r[0] for r in rows} == {"X-666"}


def test_export_queda_en_bitacora(client, db_session, tenant):
    _seed_facturas(db_session, tenant)
    client.get("/v1/export/facturas.xlsx?q=beto&bucket=por_vencer", headers=HEADERS)
    entry = db_session.scalar(
        select(AuditLog).where(
            AuditLog.tenant_id == tenant.id, AuditLog.action == "data.export"
        )
    )
    assert entry is not None
    assert entry.entity_type == "facturas"
    assert entry.after["formato"] == "xlsx"
    assert entry.after["filas"] == 1
    assert entry.after["filtros"] == {"bucket": "por_vencer", "q": "beto"}


def test_entidad_desconocida_404(client, db_session, tenant):
    res = client.get("/v1/export/recordatorios.xlsx", headers=HEADERS)
    assert res.status_code == 404


# --- Clientes: filtros y columnas reimportables --------------------------------


def test_clientes_xlsx_filtros_q_y_tag(client, db_session, tenant):
    _cust(
        db_session, tenant, "Papelería Ana", "5215511110001",
        email="ana@x.com", tags=["tg1"], meta={"empresa": "Ana SA"},
    )
    _cust(db_session, tenant, "Ferretería Beto", "5215511110002")
    _cust(db_session, tenant, "Lead Carla", None, kind="prospecto", meta={"origen": "Feria"})

    res = client.get("/v1/export/clientes.xlsx", headers=HEADERS)
    ws, headers, rows = _sheet(res)
    assert ws.title == "clientes"
    assert headers == [
        "nombre", "telefono", "correo", "empresa", "etiquetas",
        "facturas_abiertas", "saldo_abierto",
    ]
    # Solo clientes: el prospecto no se mezcla.
    assert {r[0] for r in rows} == {"Papelería Ana", "Ferretería Beto"}
    ana = {r[0]: r for r in rows}["Papelería Ana"]
    assert ana[1] == "5215511110001" and ana[2] == "ana@x.com"
    assert ana[3] == "Ana SA" and ana[4] == "Mayoreo"

    _, _, rows = _sheet(client.get("/v1/export/clientes.xlsx?tag=tg1", headers=HEADERS))
    assert [r[0] for r in rows] == ["Papelería Ana"]
    _, _, rows = _sheet(client.get("/v1/export/clientes.xlsx?q=beto", headers=HEADERS))
    assert [r[0] for r in rows] == ["Ferretería Beto"]

    # Prospectos: sus columnas llevan origen (reimportables por smart_import).
    _, headers, rows = _sheet(client.get("/v1/export/prospectos.xlsx", headers=HEADERS))
    assert headers == ["nombre", "telefono", "correo", "empresa", "origen", "etiquetas"]
    assert [r[0] for r in rows] == ["Lead Carla"]
    assert rows[0][4] == "Feria"


# --- Las demás listas responden con su forma -----------------------------------


def test_productos_citas_promesas_conciliacion(client, db_session, tenant):
    db_session.add(
        Product(tenant_id=tenant.id, name="Anillo oro", sku="A-1", price=4500, stock=3, unit="pieza")
    )
    db_session.add(
        Appointment(
            tenant_id=tenant.id, title="Valuación", customer_name="Ana",
            starts_at=datetime(2026, 7, 20, 10, 30),
        )
    )
    ana = _cust(db_session, tenant, "Papelería Ana", "5215511110001")
    inv = _inv(db_session, tenant, ana, "F-010", 1200, HOY + timedelta(days=3))
    db_session.add(
        PaymentPromise(tenant_id=tenant.id, invoice_id=inv.id, promised_date=HOY, note="viernes")
    )
    db_session.add(
        Payment(
            tenant_id=tenant.id, amount=1200, currency="MXN", paid_at=HOY,
            source="banco", status="conciliado",
            meta={"aplicaciones": [{"folio": "F-010", "aplicado": 1200.0}]},
        )
    )
    db_session.flush()

    _, headers, rows = _sheet(client.get("/v1/export/productos.xlsx", headers=HEADERS))
    assert headers == ["nombre", "sku", "precio", "existencia", "unidad", "procedencia"]
    assert rows[0][0] == "Anillo oro" and rows[0][2] == 4500.0

    _, headers, rows = _sheet(client.get("/v1/export/citas.xlsx", headers=HEADERS))
    assert headers == ["titulo", "cliente", "telefono", "fecha", "notas", "procedencia"]
    assert rows[0][0] == "Valuación" and rows[0][3] == datetime(2026, 7, 20, 10, 30)

    _, headers, rows = _sheet(client.get("/v1/export/promesas.xlsx", headers=HEADERS))
    assert headers == [
        "cliente", "folio", "monto", "fecha_promesa", "nota", "estado", "cumplida_el",
    ]
    assert rows[0][1] == "F-010" and rows[0][5] == "activa"
    # status=fulfilled: sin cumplidas todavía, solo encabezado.
    _, _, rows = _sheet(client.get("/v1/export/promesas.xlsx?status=fulfilled", headers=HEADERS))
    assert rows == []

    _, headers, rows = _sheet(client.get("/v1/export/conciliacion.xlsx", headers=HEADERS))
    assert headers == [
        "fecha_pago", "monto", "moneda", "origen", "referencia", "deposito_de",
        "estado", "facturas_aplicadas", "excedente", "resuelto_el",
    ]
    assert rows[0][1] == 1200.0 and rows[0][6] == "conciliado"
    assert rows[0][7] == "F-010"


def test_export_vacio_sirve_de_plantilla(client, db_session, tenant):
    """Sin datos, el archivo baja igual: encabezados reimportables, cero filas."""
    _, headers, rows = _sheet(client.get("/v1/export/facturas.xlsx", headers=HEADERS))
    assert headers == FACTURAS_HDR and rows == []


# --- El test redondo: export → importador → las mismas facturas -----------------


def test_roundtrip_export_import_facturas(client, db_session, tenant, otro_tenant):
    _seed_facturas(db_session, tenant)
    res = client.get("/v1/export/facturas.xlsx", headers=HEADERS)
    content = res.content
    filename = f"facturas-{HOY.isoformat()}.xlsx"

    # El importador lee el archivo tal cual y ve TODOS sus campos por nombre exacto.
    headers, rows = read_table(content, filename)
    assert headers == FACTURAS_HDR
    assert len(rows) == 2
    assert set(ENTITY_FIELDS["facturas"]) <= set(headers)

    # Mapeo identidad (encabezado == campo): en un negocio limpio se crean idénticas.
    mapping = {campo: campo for campo in ENTITY_FIELDS["facturas"]}
    report = commit(db_session, otro_tenant.id, content, filename, "facturas", mapping, [])
    assert report.errors == []
    assert report.created == 2 and report.skipped == 0
    importadas = {
        inv.folio: inv
        for inv in db_session.scalars(
            select(Invoice).where(Invoice.tenant_id == otro_tenant.id)
        ).all()
    }
    original = db_session.scalar(
        select(Invoice).where(Invoice.tenant_id == tenant.id, Invoice.folio == "F-001")
    )
    copia = importadas["F-001"]
    assert float(copia.amount) == float(original.amount) == 1500.50
    assert copia.due_date == original.due_date
    assert copia.issued_date == original.issued_date
    cliente_copia = db_session.get(Customer, copia.customer_id)
    assert cliente_copia.name == "Papelería Ana"
    assert cliente_copia.phone == "5215511110001"  # el teléfono sobrevive el viaje

    # De regreso al negocio de origen: se RECONOCEN como las mismas (cero duplicados).
    report2 = commit(db_session, tenant.id, content, filename, "facturas", mapping, [])
    assert report2.created == 0 and report2.skipped == 2
