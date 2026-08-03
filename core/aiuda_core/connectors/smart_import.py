"""Importador universal: subes CUALQUIER Excel y aiuda entiende qué es.

El usuario sube su archivo tal como lo lleva (Excel o CSV, con sus propios nombres
de columna). La IA hace dos cosas:
  1. CLASIFICA la hoja: ¿son facturas, clientes, productos, citas o prospectos?
  2. MAPEA las columnas del usuario a los campos de ese tipo.
Luego cada tipo se carga a su lugar (cartera, directorio, catálogo, agenda...) y
alimenta al aiudante que lo necesita. Sin plantillas, sin reacomodar nada.
"""

import csv
import io
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from aiuda_core.engine.llm import parse_json_block
from aiuda_core.engine.runner import ProviderRunner, make_runner
from aiuda_core.models import Appointment, Customer, Invoice, Product
from aiuda_core.phones import normalize_mx

# Campos destino por tipo de hoja. La IA mapea las columnas del usuario a estos.
ENTITY_FIELDS: dict[str, dict[str, str]] = {
    "facturas": {
        "folio": "identificador o número de la factura",
        "cliente": "nombre del cliente o razón social",
        "telefono": "teléfono o WhatsApp del cliente",
        "monto": "importe o total de la factura",
        "fecha_emision": "fecha de emisión o de la factura",
        "fecha_vencimiento": "fecha de vencimiento o límite de pago",
    },
    "clientes": {
        "nombre": "nombre del cliente, persona o razón social",
        "telefono": "teléfono o WhatsApp",
        "correo": "correo electrónico",
        "empresa": "empresa o negocio al que pertenece",
    },
    "prospectos": {
        "nombre": "nombre del prospecto o contacto",
        "telefono": "teléfono o WhatsApp",
        "correo": "correo electrónico",
        "empresa": "empresa del prospecto",
        "origen": "de dónde salió el prospecto (campaña, referido, evento)",
    },
    "productos": {
        "nombre": "nombre o descripción del producto",
        "sku": "clave, SKU o código del producto",
        "precio": "precio de venta",
        "existencia": "cantidad en inventario o stock",
        "unidad": "unidad de medida (pieza, kg, caja)",
    },
    "citas": {
        "titulo": "asunto, servicio o motivo de la cita",
        "cliente": "nombre del cliente de la cita",
        "telefono": "teléfono o WhatsApp del cliente",
        "fecha": "fecha y hora de la cita",
        "notas": "notas o detalles adicionales",
    },
}

ENTITY_LABEL: dict[str, str] = {
    "facturas": "Facturas por cobrar",
    "clientes": "Clientes",
    "prospectos": "Prospectos",
    "productos": "Productos",
    "citas": "Citas",
}

CLASSIFY_PROMPT = """\
Un negocio mexicano subió una hoja de cálculo. Por sus columnas y una muestra de
filas, decide qué tipo de datos contiene.

Columnas: {headers}

Muestra:
{sample}

Tipos posibles:
- facturas: documentos por cobrar (folio, monto, fecha de vencimiento)
- clientes: directorio de clientes que YA son clientes (nombre, teléfono, correo)
- prospectos: posibles clientes por contactar / leads / base para prospectar
- productos: catálogo de lo que vende (precio, existencia, SKU)
- citas: agenda de citas o eventos con fecha/hora
- desconocido: si no encaja con claridad en ninguno

Responde ÚNICAMENTE un JSON: {{"tipo": "...", "confianza": 0.0}}
"""

MAPPING_PROMPT = """\
Un negocio mexicano subió una hoja de tipo "{entity}". Estas son sus columnas y una
muestra de filas:

Columnas: {headers}

Muestra:
{sample}

Identifica qué columna del archivo corresponde a cada campo destino:
{targets}

Responde ÚNICAMENTE un objeto JSON donde la llave es el campo destino y el valor es
el nombre EXACTO de la columna del archivo (o null si no existe).
"""


@dataclass
class ImportReport:
    entity: str = ""
    entity_label: str = ""
    confidence: float = 0.0
    mapping: dict[str, str | None] = field(default_factory=dict)
    created: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


# --- Lectura del archivo ----------------------------------------------------


def _dedupe_headers(raw: list) -> list[str]:
    """Renombra encabezados repetidos para no perder columnas: Excel/CSV sí permiten
    dos 'Empresa', pero un dict los colapsa (gana el último). 'Empresa','Empresa' ->
    'Empresa','Empresa (2)'. Los vacíos quedan en '' y se filtran después."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in raw:
        name = (str(h).strip() if h is not None else "")
        if not name:
            out.append("")
            continue
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        out.append(name)
    return out


def read_table(content: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Lee CSV o XLSX y devuelve (headers, filas como dicts). Mapea por posición
    contra encabezados de-duplicados: ninguna columna se pierde por nombre repetido."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = _dedupe_headers(list(next(rows_iter)))
        rows = []
        for raw in rows_iter:
            row = {h: raw[i] if i < len(raw) else None for i, h in enumerate(headers) if h}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return [h for h in headers if h], rows
    # CSV (con BOM-tolerancia); csv.reader + posición, no DictReader (colapsa repetidos)
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        headers = _dedupe_headers(next(reader))
    except StopIteration:
        return [], []
    rows = []
    for raw in reader:
        row = {h: (raw[i] if i < len(raw) else None) for i, h in enumerate(headers) if h}
        if any((str(v).strip() if v is not None else "") for v in row.values()):
            rows.append(row)
    return [h for h in headers if h], rows


def _sample(headers: list[str], rows: list[dict]) -> str:
    return "\n".join(
        json.dumps({h: str(row.get(h, ""))[:40] for h in headers}, ensure_ascii=False)
        for row in rows[:3]
    )


# --- Clasificación y mapeo (IA) ---------------------------------------------


def classify_sheet(
    headers: list[str], rows: list[dict], runner: ProviderRunner | None = None
) -> tuple[str, float]:
    """La IA decide qué tipo de datos trae la hoja."""
    runner = runner or make_runner(None)
    raw = runner.complete(
        system="Clasificas hojas de cálculo de PyMEs mexicanas. Respondes solo JSON.",
        user=CLASSIFY_PROMPT.format(headers=headers, sample=_sample(headers, rows)),
        role="triage",
        task="clasificar_archivo",
        max_tokens=80,
    )
    data = parse_json_block(raw) or {}
    tipo = str(data.get("tipo", "desconocido")).strip().lower()
    if tipo not in ENTITY_FIELDS:
        tipo = "desconocido"
    try:
        conf = float(data.get("confianza", 0.0))
    except (TypeError, ValueError):
        conf = 0.0
    return tipo, conf


def infer_mapping(
    entity: str, headers: list[str], rows: list[dict], runner: ProviderRunner | None = None
) -> dict[str, str | None]:
    """La IA mapea las columnas del usuario a los campos del tipo detectado."""
    runner = runner or make_runner(None)
    fields = ENTITY_FIELDS[entity]
    targets = "\n".join(f"- {k}: {v}" for k, v in fields.items())
    raw = runner.complete(
        system="Eres experto en datos de PyMEs mexicanas. Respondes solo JSON.",
        user=MAPPING_PROMPT.format(
            entity=entity, headers=headers, sample=_sample(headers, rows), targets=targets
        ),
        role="triage",
        task="mapear_archivo",
        max_tokens=300,
    )
    mapping = parse_json_block(raw) or {}
    return {k: (mapping.get(k) if mapping.get(k) in headers else None) for k in fields}


# --- Parsers tolerantes -----------------------------------------------------


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    raise ValueError(f"fecha no reconocida: {value!r}")


def _parse_datetime(value) -> datetime | None:
    """Fecha con hora si la trae; si solo hay fecha, queda a medianoche."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M",
        "%d/%m/%Y %H:%M",
        "%d/%m/%y %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text[:16], fmt)
        except ValueError:
            continue
    try:
        d = _parse_date(value)
        return datetime(d.year, d.month, d.day)
    except ValueError:
        return None


def _parse_amount(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value).replace(",", ""))
    return float(cleaned)


def _parse_amount_opt(value) -> float | None:
    try:
        if value in (None, ""):
            return None
        return _parse_amount(value)
    except (ValueError, TypeError):
        return None


def _clean_phone(value) -> str:
    # Regla compartida con el envío (wacli/Evolution): ver aiuda_core.phones.
    return normalize_mx(value)


def _getter(mapping: dict[str, str | None]):
    def get(row: dict, fieldname: str):
        col = mapping.get(fieldname)
        return row.get(col) if col else None

    return get


def _row_extras(row: dict, extras) -> dict:
    """Columnas que el usuario marcó como 'dato extra': se guardan tal cual, con
    el nombre de su columna como llave. No se pierde nada del Excel original."""
    out = {}
    for col in extras or ():
        v = row.get(col)
        if v not in (None, ""):
            out[col] = str(v).strip()
    return out


def _cell(value) -> str:
    if value is None:
        return ""
    return str(value)[:48]


# --- Ingesta por tipo -------------------------------------------------------


def _ingest_facturas(session, tenant_id, rows, mapping, extras=(), origin=None):
    """Cartera. Idempotente por (tenant, folio): re-subir no duplica."""
    from aiuda_core.engine.presence import add_presence

    get = _getter(mapping)
    created = skipped = 0
    errors: list[str] = []
    required = ("folio", "cliente", "monto", "fecha_vencimiento")
    missing = [f for f in required if not mapping.get(f)]
    if missing:
        return 0, 0, [f"No identifiqué columnas para: {', '.join(missing)}"]
    for i, row in enumerate(rows, start=2):
        try:
            folio = str(get(row, "folio") or "").strip()
            if not folio or folio == "None":
                skipped += 1
                continue
            exists = session.scalar(
                select(Invoice).where(Invoice.tenant_id == tenant_id, Invoice.folio == folio)
            )
            if exists:
                add_presence(exists, "excel", folio)
                skipped += 1
                continue
            name = str(get(row, "cliente") or "").strip()
            phone = _clean_phone(get(row, "telefono")) or "0000000000"
            customer = session.scalar(
                select(Customer).where(Customer.tenant_id == tenant_id, Customer.phone == phone)
            )
            if customer is None or phone == "0000000000":
                customer = session.scalar(
                    select(Customer).where(Customer.tenant_id == tenant_id, Customer.name == name)
                )
            if customer is None:
                customer = Customer(tenant_id=tenant_id, name=name, phone=phone)
                session.add(customer)
                session.flush()
            due = _parse_date(get(row, "fecha_vencimiento"))
            issued_raw = get(row, "fecha_emision")
            issued = _parse_date(issued_raw) if issued_raw not in (None, "") else due
            session.add(
                Invoice(
                    tenant_id=tenant_id,
                    customer_id=customer.id,
                    folio=folio,
                    amount=_parse_amount(get(row, "monto")),
                    issued_date=issued,
                    due_date=due,
                    source="excel",
                    presence={"excel": {"ref": folio, **(origin or {})}},
                    meta=_row_extras(row, extras),
                )
            )
            created += 1
        except (ValueError, TypeError, AttributeError) as exc:
            errors.append(f"Fila {i}: {exc}")
    session.flush()
    return created, skipped, errors


def _ingest_personas(session, tenant_id, rows, mapping, kind, extras=(), origin=None):
    """Clientes y prospectos. Upsert por (tenant, teléfono) cuando hay teléfono;
    si no, deduplica por (tenant, nombre). Un cliente sin teléfono NO se pierde:
    se carga marcado 'sin contacto' (aún no se le puede escribir por WhatsApp)."""
    get = _getter(mapping)
    created = skipped = sin_tel = 0
    errors: list[str] = []
    for i, row in enumerate(rows, start=2):
        try:
            name = str(get(row, "nombre") or "").strip()
            if not name:
                skipped += 1
                continue
            phone = _clean_phone(get(row, "telefono")) or None
            email = (str(get(row, "correo") or "").strip()) or None
            meta = _row_extras(row, extras)
            empresa = str(get(row, "empresa") or "").strip()
            if empresa:
                meta["empresa"] = empresa
            origen = str(get(row, "origen") or "").strip()
            if origen:
                meta["origen"] = origen
            if phone:
                existing = session.scalar(
                    select(Customer).where(
                        Customer.tenant_id == tenant_id, Customer.phone == phone
                    )
                )
            else:
                # Sin teléfono: deduplica por nombre para no duplicar en re-cargas.
                sin_tel += 1
                existing = session.scalar(
                    select(Customer).where(
                        Customer.tenant_id == tenant_id,
                        Customer.phone.is_(None),
                        func.lower(Customer.name) == name.lower(),
                    )
                )
            if existing:
                if email and not existing.email:
                    existing.email = email
                if meta:
                    existing.meta = {**(existing.meta or {}), **meta}
                if origin:
                    existing.presence = {**(existing.presence or {}), "excel": dict(origin)}
                # No degradar un cliente a prospecto; sí ascender prospecto->cliente.
                if kind == "cliente":
                    existing.kind = "cliente"
                skipped += 1
            else:
                session.add(
                    Customer(
                        tenant_id=tenant_id,
                        name=name,
                        phone=phone,
                        email=email,
                        kind=kind,
                        meta=meta,
                        presence={"excel": dict(origin)} if origin else {},
                    )
                )
                created += 1
        except (ValueError, TypeError, AttributeError) as exc:
            errors.append(f"Fila {i}: {exc}")
    session.flush()
    if sin_tel:
        # Informativo, no es un error: se cargaron, solo falta su teléfono.
        errors.append(
            f"{sin_tel} sin teléfono: cargados, pero aún no se les puede escribir por "
            "WhatsApp (agrega su teléfono cuando lo tengas)."
        )
    return created, skipped, errors


def _ingest_productos(session, tenant_id, rows, mapping, extras=(), origin=None):
    """Catálogo. Upsert por SKU si lo hay, si no por nombre."""
    get = _getter(mapping)
    created = skipped = 0
    errors: list[str] = []
    for i, row in enumerate(rows, start=2):
        try:
            name = str(get(row, "nombre") or "").strip()
            if not name:
                skipped += 1
                continue
            sku = (str(get(row, "sku") or "").strip()) or None
            price = _parse_amount_opt(get(row, "precio"))
            stock = _parse_amount_opt(get(row, "existencia"))
            unit = (str(get(row, "unidad") or "").strip()) or None
            existing = None
            if sku:
                existing = session.scalar(
                    select(Product).where(Product.tenant_id == tenant_id, Product.sku == sku)
                )
            if existing is None:
                existing = session.scalar(
                    select(Product).where(Product.tenant_id == tenant_id, Product.name == name)
                )
            if existing:
                if price is not None:
                    existing.price = price
                if stock is not None:
                    existing.stock = stock
                skipped += 1
            else:
                session.add(
                    Product(
                        tenant_id=tenant_id,
                        name=name,
                        sku=sku,
                        price=price,
                        stock=stock,
                        unit=unit,
                        source="excel",
                        presence={"excel": {"ref": sku or name, **(origin or {})}},
                        meta=_row_extras(row, extras),
                    )
                )
                created += 1
        except (ValueError, TypeError, AttributeError) as exc:
            errors.append(f"Fila {i}: {exc}")
    session.flush()
    return created, skipped, errors


def _ingest_citas(session, tenant_id, rows, mapping, extras=(), origin=None):
    """Agenda. Dedup por (título, fecha/hora). (La cita no tiene UI de
    procedencia hoy; el origin se ignora aquí.)"""
    get = _getter(mapping)
    created = skipped = 0
    errors: list[str] = []
    for i, row in enumerate(rows, start=2):
        try:
            title = str(get(row, "titulo") or "").strip()
            if not title:
                skipped += 1
                continue
            starts = _parse_datetime(get(row, "fecha"))
            cust_name = (str(get(row, "cliente") or "").strip()) or None
            phone = _clean_phone(get(row, "telefono")) or None
            notes = (str(get(row, "notas") or "").strip()) or None
            existing = session.scalar(
                select(Appointment).where(
                    Appointment.tenant_id == tenant_id,
                    Appointment.title == title,
                    Appointment.starts_at == starts,
                )
            )
            if existing:
                skipped += 1
                continue
            session.add(
                Appointment(
                    tenant_id=tenant_id,
                    title=title,
                    customer_name=cust_name,
                    customer_phone=phone,
                    starts_at=starts,
                    notes=notes,
                    source="excel",
                    meta=_row_extras(row, extras),
                )
            )
            created += 1
        except (ValueError, TypeError, AttributeError) as exc:
            errors.append(f"Fila {i}: {exc}")
    session.flush()
    return created, skipped, errors


_INGEST = {
    "facturas": _ingest_facturas,
    "clientes": lambda s, t, r, m, e=(), o=None: _ingest_personas(s, t, r, m, "cliente", e, o),
    "prospectos": lambda s, t, r, m, e=(), o=None: _ingest_personas(s, t, r, m, "prospecto", e, o),
    "productos": _ingest_productos,
    "citas": _ingest_citas,
}


def _origin(filename: str, at: datetime | None) -> dict:
    """Procedencia: de qué archivo y cuándo se subió. La fecha es hora de pared
    (naive) para mostrar el día tal cual, sin reinterpretar zona horaria."""
    return {"file": filename, "at": (at or datetime.now()).replace(microsecond=0).isoformat()}


# --- Orquestador ------------------------------------------------------------


def smart_import(
    session: Session,
    tenant_id: str,
    content: bytes,
    filename: str,
    runner: ProviderRunner | None = None,
    at: datetime | None = None,
) -> ImportReport:
    """Lee el archivo, detecta qué tipo de datos trae y lo carga a su lugar."""
    runner = runner or make_runner(None)
    headers, rows = read_table(content, filename)
    if not headers or not rows:
        return ImportReport(errors=["El archivo viene vacío."])
    entity, confidence = classify_sheet(headers, rows, runner)
    report = ImportReport(
        entity=entity,
        entity_label=ENTITY_LABEL.get(entity, "Desconocido"),
        confidence=confidence,
    )
    if entity not in ENTITY_FIELDS:
        report.errors.append(
            "No reconocí qué tipo de datos trae esta hoja. Asegúrate de que sean "
            "facturas, clientes, productos, citas o prospectos."
        )
        return report
    mapping = infer_mapping(entity, headers, rows, runner)
    report.mapping = mapping
    origin = _origin(filename, at)
    created, skipped, errors = _INGEST[entity](session, tenant_id, rows, mapping, (), origin)
    report.created, report.skipped, report.errors = created, skipped, errors
    return report


# --- Uploader con mapeo (dos pasos: analizar -> mapear/confirmar) ------------


def analyze(
    content: bytes,
    filename: str,
    runner: ProviderRunner | None = None,
    entity: str | None = None,
) -> dict:
    """Paso 1: lee el archivo y propone tipo + mapeo SIN importar nada. Si llega
    `entity`, no clasifica: usa ese tipo (cuando el usuario lo cambia a mano)."""
    runner = runner or make_runner(None)
    headers, rows = read_table(content, filename)
    if not headers or not rows:
        return {
            "entity": "",
            "confidence": 0.0,
            "columns": [],
            "sample": [],
            "mapping": {},
            "fields": {},
            "row_count": 0,
        }
    if entity in ENTITY_FIELDS:
        confidence = 1.0
    else:
        entity, confidence = classify_sheet(headers, rows, runner)
    mapping: dict[str, str | None] = {}
    fields: dict[str, str] = {}
    if entity in ENTITY_FIELDS:
        mapping = infer_mapping(entity, headers, rows, runner)
        fields = ENTITY_FIELDS[entity]
    sample = [{h: _cell(row.get(h)) for h in headers} for row in rows[:3]]
    return {
        "entity": entity if entity in ENTITY_FIELDS else "",
        "confidence": confidence,
        "columns": headers,
        "sample": sample,
        "mapping": mapping,
        "fields": fields,
        "row_count": len(rows),
    }


def commit(
    session: Session,
    tenant_id: str,
    content: bytes,
    filename: str,
    entity: str,
    mapping: dict,
    extras,
    at: datetime | None = None,
) -> ImportReport:
    """Paso 2: importa con el mapeo que confirmó el usuario. Las columnas no
    mapeadas marcadas como extra se guardan; lo demás se ignora."""
    if entity not in ENTITY_FIELDS:
        return ImportReport(errors=["Elige un tipo válido para importar."])
    headers, rows = read_table(content, filename)
    if not rows:
        return ImportReport(errors=["El archivo viene vacío."])
    clean = {
        k: v
        for k, v in (mapping or {}).items()
        if k in ENTITY_FIELDS[entity] and v in headers
    }
    # Sin mapeo utilizable no se importa NADA, y hasta ahora eso salía como
    # "created: 0, skipped: N, errors: []": el dueño subía su Excel, no entraba una
    # sola fila y la consola no le decía por qué. Un no-op silencioso en el camino
    # más publicitado del producto. Ahora se dice, y se dice en su idioma.
    if not clean:
        columnas = ", ".join(headers[:6]) or "ninguna"
        return ImportReport(
            entity=entity,
            entity_label=ENTITY_LABEL.get(entity, ""),
            skipped=len(rows),
            errors=[
                f"No se importó nada porque no quedó claro qué columna es cuál. "
                f"El archivo trae: {columnas}. Dinos qué columna corresponde a "
                f"{', '.join(ENTITY_FIELDS[entity])} y lo cargamos."
            ],
        )
    used = set(clean.values())
    extra_cols = [c for c in (extras or []) if c in headers and c not in used]
    report = ImportReport(
        entity=entity, entity_label=ENTITY_LABEL.get(entity, ""), mapping=clean
    )
    origin = _origin(filename, at)
    created, skipped, errors = _INGEST[entity](
        session, tenant_id, rows, clean, extra_cols, origin
    )
    report.created, report.skipped, report.errors = created, skipped, errors
    # Red de seguridad: "no entró nada y no hay error" no es una respuesta. Si el
    # ingestor se saltó todo sin explicar, se explica aquí antes de contestarle al
    # dueño, que lo único que ve es que su archivo no hizo nada.
    if created == 0 and skipped > 0 and not errors:
        faltantes = [c for c in ENTITY_FIELDS[entity] if c not in clean]
        detalle = (
            f" Puede que falte decir cuál columna es {', '.join(faltantes[:3])}."
            if faltantes
            else " Puede que ya estuvieran cargados."
        )
        report.errors = [
            f"No se cargó ninguno de los {skipped} renglones del archivo.{detalle}"
        ]
    return report
